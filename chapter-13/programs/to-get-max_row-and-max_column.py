import openpyxl
sheet = openpyxl.load_workbook('Book1.xlsx')
select_sheet = sheet['Sheet1']
r = select_sheet.max_row
print("Total Rows",r)
c = select_sheet.max_column
print("Total column",c)
