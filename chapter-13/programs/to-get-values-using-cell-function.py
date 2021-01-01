import openpyxl
sheet = openpyxl.load_workbook('Book1.xlsx')
select_sheet = sheet['Sheet1']
item = select_sheet.cell(row=2, column=2).value
print(item)
for i in range(1,5):
    print(select_sheet.cell(row=i, column=2).value)
