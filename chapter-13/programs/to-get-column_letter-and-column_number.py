import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
sheet = openpyxl.load_workbook('Book1.xlsx')
select_sheet = sheet['Sheet1']
a = get_column_letter(2)
print(a)
b = column_index_from_string('A')
print(b)
